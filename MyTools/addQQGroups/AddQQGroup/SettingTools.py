from openpyxl import load_workbook
from openpyxl import Workbook
import uuid
import os


class SettingTools:

    @staticmethod
    def check_setting_file(file_path):
        """
        检查配置文件是否存在，存在则返回[[urls],[datetime]]，不存在则新建空配置文件，返回空值
        :return:
        """
        try:
            wb = load_workbook(file_path)
            return [[val.value for val in wb['Sheet']['A'][1:]], [val.value for val in wb['Sheet']['B'][1:]]]
        except FileNotFoundError:

            # 定义文件夹路径
            folder_path = 'D:/AddQQGroup'
            os.makedirs(folder_path, exist_ok=True)
            excel_file_path = os.path.join(folder_path, 'settings.xlsx')
            wb = Workbook()
            sheet = wb.active
            sheet.column_dimensions['A'].width = 30
            sheet.column_dimensions['B'].width = 30
            sheet['A1'] = 'QQ群聊链接'
            sheet['B1'] = '加群时间'
            sheet['C3'] = '请在<QQ群聊链接>中填写群聊加群链接，在<加群时间>中填写开始入群时间，时间格式为：2023/09/08/08/03/01'
            sheet2 = wb.create_sheet()
            sheet2.title = 'Sheet2'
            sheet2.column_dimensions['A'].width = 30
            sheet2.column_dimensions['B'].width = 50
            sheet2.column_dimensions['C'].width = 30
            sheet2['A1'] = 'QQ路径'
            sheet2['B1'] = '是否启动时直接执行(填 是/否)'
            sheet2['B2'] = '否'
            sheet2['C1'] = '登陆码'
            sheet2['C2'] = SettingTools.get_mac()
            wb.save(filename=excel_file_path)
            wb.close()
            return None

    @staticmethod
    def find_file_path(file_name):
        """
        查找指定文件的绝对路径
        :param file_name: 文件名称
        :return:
        """
        for root, dirs, files in os.walk('/'):
            if file_name in files:
                return os.path.join(root, file_name)
        return None

    @staticmethod
    def check_qq_path(file_path):
        """
        检查配置文件中是否有QQ路径，若无，则加载路径,加载失败返回加载信息
        :param file_path: 配置文件路径
        :return: 未找到（报错消息）
        """
        workbook = load_workbook(file_path)
        worksheet = workbook['Sheet2']
        if worksheet['A2'].value is None:
            qq_path = SettingTools.find_file_path('QQScLauncher.exe')
            if qq_path is not None:
                worksheet['A2'].value = qq_path
                workbook.save(file_path)
                workbook.close()
                return None
            else:
                return "未找到QQ路径，请检查拿您的电脑是否已安装QQ！"

    @staticmethod
    def open_setting_file(file_path):
        """
        运行指定可执行文件
        :param file_path: 可执行文件路径
        :return:
        """
        os.startfile(file_path)

    @staticmethod
    def get_mac():
        """
        获取mac地址
        :return:
        """
        mac = uuid.UUID(int=uuid.getnode()).hex[-12:]
        return "".join([mac[e:e + 2] for e in range(0, 11, 2)])


if __name__ == "__main__":
    # f_path = r'D:/AddQQGroup/settings.xlsx'
    # a = SettingTools.check_setting_file(f_path)
    # print(a)
    print(SettingTools.get_mac())
